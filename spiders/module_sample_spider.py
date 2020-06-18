import sys
import traceback
from scrapy.crawler import CrawlerProcess
import scrapy
import logging
logging.getLogger('scrapy').setLevel(logging.WARNING)
logging.getLogger('scrapy').propagate = False
from multiprocessing import Process, Queue

DATAS = {}

class DetailPage(scrapy.Spider):
    
    name = "detail_spider"
    
    start_urls = [""]
    q = None
    def parse(self, response):
        print('\n********** Second Status Info **********\n')
        if response.status == 200:
            ContactDatas = self.q.get()
            print (f"URL :\t\t{response.url} ... OK\n")
            section = response.xpath("//article[@id='listing-title']")
            contacts = section.css("div.contacts-list").css("div.contact-row").css("div.contact-item")
            contactsDatas = []
            for j in range(len(contacts)):
                data = {}
                name = contacts[j].css("strong::text")
                jobs = contacts[j].css("p::text")
                phone = contacts[j].css("a.phone-data::text")
                data["Contact Name"] = name.extract()
                data["Contact Jobs"] = jobs.extract()
                data["Contact Phone"] = phone.extract()
                print (f"Name :\t{name.extract()}\tJob:\t{jobs.extract()}\tPhone:\t{phone.extract()}")
                contactsDatas.append(data)
            ContactDatas[self.start_urls[0]]["ContactDatas"] = contactsDatas
            self.q.put(ContactDatas)
        print('\n********** Second Status Info **********\n')

    
def run_detail_spider(url, Datas):
	process = CrawlerProcess({'USER_AGENT': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36'})
	process.crawl(DetailPage, start_urls = [url,], q = Datas)
	process.start()

class module_sample_spider(scrapy.Spider):

    name = "sample_spider"

    start_urls = ['https://directory.retailcouncil.org/search-results/?c=all&r=all#search-results', ]
    q = None
    def parse(self, response):
        print('\n********** First Status Info **********\n')
        if response.status == 200:
            CompanyDatas = {}
            print (f"URL :\t\t{response.url} ... OK\n")
            
            section = response.xpath("//section[@class='search-results']//div[@class='results']//section[@class='featured-listing premier-results']")

            items = section.css("div.listing-item")
            
            for i in range(len(items)):
                data = {}
                item = items[i].css("div.listing-meta")
                # print (item.css("div.listing-meta").css("h4").css("a::text").extract_first())
                company_name = item.css("h4").css("a::text")
                # print (f"Company Name :\t{company_name.extract_first()}")
                details = item.css("div.col-md-7")
                address = details.css("p::text")
                # print (f"Address :\t{''.join(address.extract()[:3])}")
                phone_websit = details.css("div.contact")
                phone = phone_websit.css("a.phone-data::text")
                # print (f"Phone :\t\t{phone.extract_first()}")
                websit = phone_websit.css("a")[-1].css("::text")
                # print (f"Websit :\t{websit.extract_first()}")
                NEXT_WEB = items[i].css("div.row").css("div.button-wrap").css("a::attr('href')").extract_first()
                # print (f"Next Page :\t{NEXT_WEB}\n")
                data["Company Name"] = company_name.extract_first()
                data["Company Address"] = ''.join(address.extract()[:3])
                data["Company Phone"] = phone.extract_first()
                data["Company Websit"] = websit.extract_first()
                data["Contacts Page"] = NEXT_WEB
                CompanyDatas[NEXT_WEB] = {"CompanyData":data}
            self.q.put(CompanyDatas)
        print('\n********** First Status Info **********\n')
        return
    
def run_module(Datas):
    process = CrawlerProcess({'USER_AGENT': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36'})
    process.crawl(module_sample_spider, q= Datas)
    process.start()
    
def create_excel(file_name, data_init):
	df = pandas.DataFrame(data_init)
	writer = pandas.ExcelWriter(file_name)
	df.to_excel(writer, index = False, header= False)
	writer.save()
	pass

def write_excel(file_name, data):
	df = pandas.DataFrame(data)
	book = load_workbook(file_name)
	writer = pandas.ExcelWriter(file_name, engine='openpyxl')
	writer.book = book
	writer.sheets = {ws.title: ws for ws in book.worksheets}
	for sheetname in writer.sheets:
		df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header= False)
	writer.save()
	pass

if __name__ == "__main__":
    Q = Queue()
    P = Process(target= run_module, args= (Q,))
    P.start()
    DATAS = Q.get()
    P.join()
    print (DATAS)
    Q.put(DATAS)
    jobs = []
    for web in DATAS:
        P = Process(target= run_detail_spider, args= (web, Q))
        P.start()
        DATAS = Q.get()
        Q.put(DATAS)
        P.join()
    DATAS = Q.get()
    for web in DATAS:
        print (web)
        print ("\tCompanyData:")
        for dataname in DATAS[web]["CompanyData"]:
            print (f"\t\t{dataname} : {DATAS[web]['CompanyData'][dataname]}")
        for contact in DATAS[web]["ContactDatas"]: # CompanyData, ContactDatas
            for contactdata in contact:
                print (f"\t\t{contactdata} : {contact[contactdata]}")
    # import json
    # with open("CompanyDatas.json", 'w', encoding= 'utf-8') as f:
    #     json.dump(DATAS,f)
    import os
    import numpy as np 
    from openpyxl import load_workbook
    import pandas
    
    Company = [['Company Name', 'Address', 'Phone', 'Website', 'Contacts', 'Phone']]
    
    datas  = list(DATAS.values())
    rows = 0
    for data in datas:
        companys = data["CompanyData"]
        name = companys["Company Name"]
        addr = companys["Company Address"]
        phone = companys["Company Phone"]
        web = companys["Company Websit"]
        Company.append([name, addr, phone, web])
        rows += 1
        contacts = data["ContactDatas"]
        numcontact = len(contacts)
        for i in range(numcontact):
            if i == 0:
                Company[rows].append(''.join(contacts[i]["Contact Name"]) + "\n\n\t" + ''.join(contacts[i]["Contact Jobs"]))
                Company[rows].append(''.join(contacts[i]["Contact Phone"]))
            else:
                Company.append(['']*4)
                rows += 1
                Company[rows].append(''.join(contacts[i]["Contact Name"]) + "\n\n\t" + ''.join(contacts[i]["Contact Jobs"]))
                Company[rows].append(''.join(contacts[i]["Contact Phone"]))
    for row in Company:
        print (row)
    FILE_PATH = "CompanyDatas.xlsx"
    # if not os.path.exists(FILE_DIR):
    #     os.mkdir(FILE_DIR)
    if not os.path.exists(FILE_PATH):
        print (f"Create and Write")
        create_excel(FILE_PATH, Company)
        #write_excel(FILE_PATH, details)
    else:
        print (f"Write")
        #write_excel(FILE_PATH, detail)

